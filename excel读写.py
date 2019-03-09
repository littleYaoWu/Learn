# from xlrd import open_workbook
# from xlutils.copy import copy
#
#
# rexcel = open_workbook("C:/Users/wuyao/Documents/python/读写.xls") # 用wlrd提供的方法读取一个excel文件
# rows = rexcel.sheets()[0].nrows # 用wlrd提供的方法获得现在已有的行数
# excel = copy(rexcel) # 用xlutils提供的copy方法将xlrd的对象转化为xlwt的对象
# table = excel.get_sheet(0) # 用xlwt对象的方法获得要操作的sheet
# values = ["1", "2", "3"]
# row = rows
# for value in values:
#     table.write(row, 3, value) # xlwt对象的写方法，参数分别是行、列、值
#     table.write(row, 4, "=SUM(E2:E5)")
#     table.write(row, 5, "lala")
#     row += 1
# excel.save("C:/Users/wuyao/Documents/python/读写.xls")




import openpyxl
import pandas as pd

df_new = pd.read_excel('C:/Users/wuyao/Documents/python/读写.xlsx')
# 按此列表的顺序进行排序
list_custom_new = ['w1', 'w2', 'w3', 'w4', 'v', 'r', 'q']
df_new['key'] = df_new['key'].astype('category')
df_new['key'].cat.set_categories(list_custom_new, inplace=True)
df_new = df_new.sort_values('key', ascending=True).reset_index(drop=True)
bool = df_new['key'].str.contains('w')
filter_data = df_new[bool]
# 插入两行空行
insertRow = pd.DataFrame([['', '', '', '', '', '', '', '', ''],['', '', '', '', '', '', '', '', '']],columns=['a', 'key', 'd', 'f', 'g', 'h', 'j', 'k', 'l'])
above = df_new.loc[:len(filter_data)-1]
below = df_new.loc[len(filter_data):]
newdata = above.append(insertRow,ignore_index=True).append(below,ignore_index=True)
newdata.to_excel('C:/Users/wuyao/Documents/python/读写.xlsx', index=False)
# 追加写入公式
wb = openpyxl.load_workbook('C:/Users/wuyao/Documents/python/读写.xlsx', data_only=True)
sheet = wb.active
gs1 = '=SUM(C2:C'+str(len(filter_data)+1)+')'
gs2 = '=SUM(D2:D'+str(len(filter_data)+1)+')'
sheet.cell(row=len(filter_data)+2, column=3, value=gs1)
sheet.cell(row=len(filter_data)+2, column=4, value=gs2)
wb.save('C:/Users/wuyao/Documents/python/读写.xlsx')


